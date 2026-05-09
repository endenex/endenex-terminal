#!/usr/bin/env python3
"""
CAISO Generator Interconnection Queue → repowering_projects.

CAISO publishes a public CSV/XLS of all active interconnection requests in
California. The file is updated weekly. We pull it, filter to wind/solar/
storage, normalise to our five-stage enum, and upsert.

URL (verified May 2026):
  http://www.caiso.com/Documents/PublicQueueReport.xlsx

CAISO queue stages (Phase column):
  - Cluster Study (annual cluster window)
  - Phase I Study (System Impact)
  - Phase II Study (Facilities)
  - GIA Negotiation (Generator Interconnection Agreement)
  - GIA Executed (interconnection agreement signed)
  - In Service / Withdrawn

The California pipeline plus standalone-storage queue contains ~250-400 GW
worth of projects, dominated by solar+storage hybrids. After Texas, this
is the next-largest US ISO pipeline.
"""

from __future__ import annotations

import argparse
import re
import sys
from io import BytesIO

import requests

from base_ingestor import get_supabase_client, log
from repowering._base import (
    normalise_stage, normalise_asset_class,
    upsert_project, today_iso, parse_date, is_too_old,
)


CAISO_URL = 'http://www.caiso.com/Documents/PublicQueueReport.xlsx'

# STRICT REPOWERING / DECOMMISSIONING FILTER (per user 2026-05-09):
# The repowering_projects table is meant exclusively for projects that
# are tearing down an existing asset and replacing it. Net-new
# greenfield projects, "EXPANSION", "PHASE 2/3/4", and hybrid additions
# do NOT belong here. CAISO's PublicQueueReport is overwhelmingly new
# greenfield development; only ~5% of rows are actual repowers.
#
# Keep a row only if its name matches an explicit repowering signal.
# Reject names with expansion / phase / hybrid markers even if the
# repower regex would otherwise match.
REPOWER_NAME_RE = re.compile(
    r'\b(repower(ing|ed)?|decommission(ing|ed)?|dismantl(e|ing|ement)|'
    r'demolition|retire(ment)?|replacement\s+of|modernization)\b',
    re.I,
)
EXPANSION_RE = re.compile(
    r'\b(expansion|expand|phase\s+(?:2|3|4|II|III|IV)|'
    r'^(?:.*\s)?(?:II|III|IV)\b|hybrid(?:ization)?)\b',
    re.I,
)


def is_strict_repowering(name: str) -> bool:
    """True only if name explicitly indicates an existing-asset repower
    or decommissioning, AND doesn't carry expansion / new-phase markers.
    """
    if not name:
        return False
    if EXPANSION_RE.search(name):
        return False
    return bool(REPOWER_NAME_RE.search(name))

# Sheet structure as of May 2026 (verified by direct inspection):
#   Sheet "Grid GenerationQueue" — active queue (~303 rows)
#   Header row is row index 3 (rows 0-2 are titles / merged-cell labels)
#   Type-1 values for renewables: 'Storage', 'Photovoltaic', 'Wind Turbine'
#   Stage signal comes from "Interconnection Agreement Status" column
SHEET_NAME = 'Grid GenerationQueue'
HEADER_ROW = 3   # 0-indexed

# Map CAISO Type-1 values to our asset_class enum
CAISO_TYPE1_MAP = {
    'storage':           'bess',
    'photovoltaic':      'solar_pv',
    'solar':             'solar_pv',
    'wind turbine':      'onshore_wind',
    'wind':              'onshore_wind',
}

# Map CAISO IA Status → our five-stage enum. IA = Interconnection
# Agreement; "Executed" means the project has signed terms with the
# utility, which is closest to our "permitted" stage. Projects still
# in study phase have IA Status = null.
CAISO_IA_STAGE_MAP = {
    'executed':           'permitted',
    'in progress':        'application_approved',
    'filed unexecuted':   'application_submitted',
}


def fetch_workbook() -> bytes:
    r = requests.get(CAISO_URL, timeout=120, headers={'User-Agent': 'endenex-terminal/1.0'})
    r.raise_for_status()
    return r.content


def parse_workbook(xls_bytes: bytes) -> list[dict]:
    try:
        import openpyxl
    except ImportError:
        log.error('openpyxl required: pip install openpyxl')
        sys.exit(1)
    wb = openpyxl.load_workbook(BytesIO(xls_bytes), data_only=True, read_only=True)
    if SHEET_NAME not in wb.sheetnames:
        log.error(f'  sheet {SHEET_NAME!r} not found; available: {wb.sheetnames}')
        return []
    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) <= HEADER_ROW:
        return []
    # Strip newlines from headers (CAISO uses \n inside header strings)
    headers = [str(h or '').replace('\n', ' ').strip() for h in rows[HEADER_ROW]]
    out: list[dict] = []
    for r in rows[HEADER_ROW + 1:]:
        if not r or all(c is None for c in r):
            continue
        out.append(dict(zip(headers, r)))
    return out


def build_row(rec: dict, today: str) -> dict | None:
    # Asset class — Type-1 carries the primary technology
    type1 = (rec.get('Type-1') or '').strip().lower()
    asset_class = CAISO_TYPE1_MAP.get(type1)
    if not asset_class:
        return None   # not wind / solar / storage

    project_name = (rec.get('Project Name') or '').strip()
    if not project_name:
        return None
    # Strict repowering filter — drop new builds, expansions, phases.
    if not is_strict_repowering(project_name):
        return None

    # Stage — prefer IA Status (executed = permitted), fall back to study
    # process if IA still null
    ia_status = (rec.get('Interconnection Agreement  Status')
                 or rec.get('Interconnection Agreement Status') or '').strip().lower()
    stage = CAISO_IA_STAGE_MAP.get(ia_status)
    if not stage:
        # Fall back to phase signal
        if rec.get('Facilities Study (FAS) or  Phase II Cluster Study'):
            stage = 'application_approved'
        elif rec.get('System Impact Study or  Phase I Cluster Study'):
            stage = 'application_submitted'
        else:
            stage = 'announced'

    queue_pos = rec.get('Queue Position')
    queue_pos = str(queue_pos).strip() if queue_pos is not None else ''

    # Net MWs to Grid is the export-side capacity; falls back to MW-1
    capacity = rec.get('Net MWs to Grid') or rec.get('MW-1')
    try:
        capacity_mw = float(capacity) if capacity else None
    except (TypeError, ValueError):
        capacity_mw = None

    county = (rec.get('County') or '').strip()
    state = (rec.get('State') or 'CA').strip()
    location = f'{county}, {state}' if county else f'{state}, USA'

    queue_date = parse_date(rec.get('Queue Date'))
    cod = parse_date(rec.get('Current On-line Date')
                     or rec.get('Proposed On-line Date (as filed with IR)'))
    # 3-year cutoff — if the queue date AND the on-line date are both
    # >3 years old, this project is stale and probably abandoned.
    if is_too_old(queue_date, today) and is_too_old(cod, today):
        return None

    return {
        'project_name':        project_name,
        'country_code':        'US',
        'asset_class':         asset_class,
        'stage':               stage,
        'stage_date':          cod or queue_date or today,
        'capacity_mw':         capacity_mw,
        'developer':           (rec.get('Interconnection Customer') or rec.get('Developer') or None),
        'operator':            None,
        'planning_reference':  queue_pos or None,
        'location_description': location,
        'source_url':          'http://www.caiso.com/PublishedDocs/Public/PublicQueueReport.xlsx',
        'notes':               f'CAISO Queue · Q# {queue_pos}' if queue_pos else 'CAISO Queue',
        'source_type':         'caiso_queue',
        'source_date':         today,
        'confidence':          'High',
        'derivation':          'Observed',
        'last_reviewed':       today,
        'external_source':     'caiso_queue',
        'external_source_id':  queue_pos or None,
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--dry-run', action='store_true')
    args = ap.parse_args()

    client = get_supabase_client()
    today = today_iso()
    log.info(f'=== CAISO queue ingestion · {today} ===')

    xls = fetch_workbook()
    log.info(f'  fetched {len(xls)/1024/1024:.1f} MB')
    rows = parse_workbook(xls)
    log.info(f'  parsed {len(rows)} rows')

    inserted = skipped = 0
    for rec in rows:
        row = build_row(rec, today)
        if not row:
            skipped += 1
            continue
        if args.dry_run:
            log.info(f'    {row["project_name"]} [{row["asset_class"]}/{row["stage"]}] · {row["capacity_mw"]} MW · {row["developer"] or "—"}')
            continue
        if upsert_project(client, row):
            inserted += 1
        else:
            skipped += 1

    if not args.dry_run:
        client.table('ingestion_runs').insert({
            'pipeline':           'caiso_queue_repowering',
            'status':             'success',
            'started_at':         f'{today}T00:00:00Z',
            'finished_at':        f'{today}T00:00:00Z',
            'records_written':    inserted,
            'source_attribution': f'CAISO Public Queue Report ({CAISO_URL})',
            'notes':              f'CAISO queue ingestion · {inserted} upserts · {skipped} skipped.',
        }).execute()

    log.info(f'=== complete: {inserted} upserted · {skipped} skipped ===')


if __name__ == '__main__':
    main()
