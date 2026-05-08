#!/usr/bin/env python3
"""
AEMO Generation Information (NEM + WEM) → repowering_projects.

AEMO (Australian Energy Market Operator) publishes a monthly XLS that
covers every existing, committed, advanced, and emerging generator on
the National Electricity Market (NEM) plus a separate workbook for the
Western Australia WEM.

URL pattern (verified May 2026):
  https://www.aemo.com.au/-/media/files/electricity/nem/planning_and_forecasting/
    generation_information/{YYYY}/generation-information-{MMM-YY}.xlsx

Useful columns from sheet "Existing Generation" / "New Developments":
  - Project Name                    → project_name
  - Owner / Project Sponsor         → developer
  - DUID (Dispatchable Unit ID)     → external_source_id
  - Fuel / Tech                     → asset_class
  - Reg Cap (MW) / Storage MWh      → capacity_mw
  - State                           → location_description
  - Status (Existing / Committed /
            Anticipated / Proposed) → stage
  - Project COD (Commercial Op Date)→ stage_date

Australia is a Tier 1 market; this single source covers ~95% of NEM
pipeline. WEM (Western Australia) is a separate workbook — not yet
ingested but TODO marker below.
"""

from __future__ import annotations

import argparse
import sys
from datetime import date
from io import BytesIO

import requests

from base_ingestor import get_supabase_client, log
from repowering._base import (
    normalise_stage, normalise_asset_class,
    upsert_project, today_iso, parse_date,
)


AEMO_INDEX_URL = 'https://www.aemo.com.au/energy-systems/electricity/national-electricity-market-nem/nem-forecasting-and-planning/forecasting-and-planning-data/generation-information'

# We can't easily auto-discover the latest URL without scraping the index page.
# The script accepts the URL as an argument; the workflow YAML can be updated
# monthly with the latest known link, OR a small scraper helper can find it.

AEMO_STAGE_MAP = {
    'existing':        'ongoing',
    'committed':       'permitted',
    'anticipated':     'application_approved',
    'proposed':        'application_submitted',
    'maturing':        'application_submitted',
    'emerging':        'announced',
    'publicly_announced': 'announced',
    'withdrawn':       None,
}


def fetch_workbook(url: str) -> bytes:
    r = requests.get(url, timeout=120, headers={'User-Agent': 'endenex-terminal/1.0'})
    r.raise_for_status()
    return r.content


def parse_workbook(xls_bytes: bytes) -> list[dict]:
    try:
        import openpyxl
    except ImportError:
        log.error('openpyxl required: pip install openpyxl')
        sys.exit(1)
    wb = openpyxl.load_workbook(BytesIO(xls_bytes), data_only=True, read_only=True)
    out: list[dict] = []
    for sheet_name in ('Existing Generation','New Developments','Existing','New'):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(h or '').strip() for h in rows[0]]
        for r in rows[1:]:
            if not r or all(c is None for c in r):
                continue
            rec = dict(zip(headers, r))
            rec['_sheet'] = sheet_name
            out.append(rec)
    return out


def build_row(rec: dict, today: str) -> dict | None:
    fuel = ((rec.get('Fuel') or rec.get('Tech') or rec.get('Technology Type')) or '').strip()
    asset_class = normalise_asset_class(fuel)
    if asset_class not in {'onshore_wind','offshore_wind','solar_pv','bess'}:
        return None

    project_name = (rec.get('Project Name') or rec.get('Station Name') or rec.get('Generator Name') or '').strip()
    if not project_name:
        return None

    # Status column varies by sheet; "Existing Generation" rows are always 'Existing'
    if rec.get('_sheet','').lower().startswith('existing'):
        stage_raw = 'existing'
    else:
        stage_raw = rec.get('Project Status') or rec.get('Status') or rec.get('Project Tracker') or 'proposed'
    stage = normalise_stage(stage_raw, AEMO_STAGE_MAP)
    if stage is None:
        return None  # withdrawn / unmapped

    duid = (rec.get('DUID') or rec.get('Dispatchable Unit ID') or '').strip()
    capacity = rec.get('Reg Cap (MW)') or rec.get('Capacity (MW)') or rec.get('MW') or rec.get('Nameplate (MW)')
    try:
        capacity_mw = float(capacity) if capacity else None
    except (TypeError, ValueError):
        capacity_mw = None

    state = (rec.get('State') or '').strip()
    cod = parse_date(rec.get('Project COD') or rec.get('Commercial Operation Date') or rec.get('FCAS')) or today

    return {
        'project_name':        project_name,
        'country_code':        'AU',
        'asset_class':         asset_class,
        'stage':               stage,
        'stage_date':          cod,
        'capacity_mw':         capacity_mw,
        'developer':           (rec.get('Owner') or rec.get('Project Sponsor') or None),
        'operator':            (rec.get('Operator') or None),
        'planning_reference':  duid or None,
        'location_description': f'{state}, Australia' if state else 'Australia',
        'source_url':          AEMO_INDEX_URL,
        'notes':               f'AEMO GI · DUID {duid}' if duid else 'AEMO GI',
        'source_type':         'aemo_giinr',
        'source_date':         today,
        'confidence':          'High',
        'derivation':          'Observed',
        'last_reviewed':       today,
        'external_source':     'aemo_giinr',
        'external_source_id':  duid or None,
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--url', required=True,
                    help='Direct URL to AEMO Generation Information workbook (XLSX). '
                         'Find latest at https://www.aemo.com.au/energy-systems/electricity/'
                         'national-electricity-market-nem/nem-forecasting-and-planning/'
                         'forecasting-and-planning-data/generation-information')
    ap.add_argument('--dry-run', action='store_true')
    args = ap.parse_args()

    client = get_supabase_client()
    today = today_iso()
    log.info(f'=== AEMO Generation Information ingestion · {today} ===')

    xls = fetch_workbook(args.url)
    log.info(f'  fetched {len(xls)/1024/1024:.1f} MB')
    rows = parse_workbook(xls)
    log.info(f'  parsed {len(rows)} rows')

    inserted = skipped = 0
    for rec in rows:
        row = build_row(rec, today)
        if not row:
            skipped += 1
            continue
        if args.dry_run:
            log.info(f'    {row["project_name"]} [{row["asset_class"]}/{row["stage"]}] · {row["capacity_mw"]} MW · {row["developer"] or "—"}')
            continue
        if upsert_project(client, row):
            inserted += 1
        else:
            skipped += 1

    if not args.dry_run:
        client.table('ingestion_runs').insert({
            'pipeline':           'aemo_gi_repowering',
            'status':             'success',
            'started_at':         f'{today}T00:00:00Z',
            'finished_at':        f'{today}T00:00:00Z',
            'records_written':    inserted,
            'source_attribution': f'AEMO Generation Information ({args.url})',
            'notes':              f'AEMO GI ingestion · {inserted} upserts · {skipped} skipped. TODO: WEM workbook (Western Australia) not yet covered.',
        }).execute()

    log.info(f'=== complete: {inserted} upserted · {skipped} skipped ===')


if __name__ == '__main__':
    main()
