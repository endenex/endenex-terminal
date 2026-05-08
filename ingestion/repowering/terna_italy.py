#!/usr/bin/env python3
"""
Terna Italy "Anagrafica Impianti" → repowering_projects.

Terna (Italy's TSO) publishes the Anagrafica Impianti — a register of all
generation plants connected to the Italian grid — at:
  https://download.terna.it/terna/Anagrafica_impianti_8d61b9f3-...xlsx

The exact URL changes monthly; latest is linked from:
  https://www.terna.it/it/sistema-elettrico/dispacciamento/anagrafica-impianti

Useful columns:
  - CODICE CENTRALE / CODICE_UP   → external_source_id
  - DENOMINAZIONE IMPIANTO        → project_name
  - FONTE                         → asset_class
  - POTENZA (MW)                  → capacity_mw
  - REGIONE                       → location_description
  - DATA INIZIO ESERCIZIO         → stage_date (commissioning date)
  - GESTORE                       → developer/operator

Note: Anagrafica is COMMISSIONED-ONLY register. For pipeline projects in
Italy (pre-permit / under construction) the parallel sources are:
  • GSE register (subsidy-tied projects)
  • Regional VIA (Valutazione Impatto Ambientale) registers
Both are scattered across 20 region-level portals — TODO scaffold.
"""

from __future__ import annotations

import argparse
import sys
from io import BytesIO

import requests

from base_ingestor import get_supabase_client, log
from repowering._base import (
    normalise_asset_class,
    upsert_project, today_iso, parse_date,
)


def fetch_workbook(url: str) -> bytes:
    r = requests.get(url, timeout=180, headers={'User-Agent': 'endenex-terminal/1.0'})
    r.raise_for_status()
    return r.content


def parse_workbook(xls_bytes: bytes) -> list[dict]:
    try:
        import openpyxl
    except ImportError:
        log.error('openpyxl required: pip install openpyxl')
        sys.exit(1)
    wb = openpyxl.load_workbook(BytesIO(xls_bytes), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h or '').strip() for h in rows[0]]
    return [dict(zip(headers, r)) for r in rows[1:] if r and any(c is not None for c in r)]


def build_row(rec: dict, today: str) -> dict | None:
    fonte = (rec.get('FONTE') or rec.get('Fonte') or '').strip().lower()
    asset_class_map = {
        'eolica':         'onshore_wind',
        'eolico':         'onshore_wind',
        'solare':          'solar_pv',
        'fotovoltaica':    'solar_pv',
        'fotovoltaico':    'solar_pv',
        'idroelettrica':   None,
        'accumulo':        'bess',
        'storage':         'bess',
    }
    asset_class = asset_class_map.get(fonte) or normalise_asset_class(fonte)
    if asset_class not in {'onshore_wind','offshore_wind','solar_pv','bess'}:
        return None

    project_name = (rec.get('DENOMINAZIONE IMPIANTO') or rec.get('Nome Impianto') or '').strip()
    if not project_name:
        return None

    cod = parse_date(rec.get('DATA INIZIO ESERCIZIO') or rec.get('Data Avviamento'))
    capacity_raw = rec.get('POTENZA (MW)') or rec.get('Potenza MW')
    try:
        capacity_mw = float(capacity_raw) if capacity_raw else None
    except (TypeError, ValueError):
        capacity_mw = None

    code = (rec.get('CODICE CENTRALE') or rec.get('CODICE_UP') or '').strip()
    region = (rec.get('REGIONE') or '').strip()

    return {
        'project_name':        project_name,
        'country_code':        'IT',
        'asset_class':         asset_class,
        'stage':               'ongoing',     # Anagrafica = commissioned only
        'stage_date':          cod or today,
        'capacity_mw':         capacity_mw,
        'developer':           (rec.get('GESTORE') or None),
        'operator':            (rec.get('GESTORE') or None),
        'planning_reference':  code or None,
        'location_description': f'{region}, Italy' if region else 'Italy',
        'source_url':          'https://www.terna.it/it/sistema-elettrico/dispacciamento/anagrafica-impianti',
        'notes':               'Terna Anagrafica Impianti',
        'source_type':         'terna_anagrafica',
        'source_date':         today,
        'confidence':          'High',
        'derivation':          'Observed',
        'last_reviewed':       today,
        'external_source':     'terna_anagrafica',
        'external_source_id':  code or None,
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--url', required=True, help='Direct URL to Terna Anagrafica XLSX')
    ap.add_argument('--dry-run', action='store_true')
    args = ap.parse_args()

    client = get_supabase_client()
    today = today_iso()
    log.info(f'=== Terna Italy ingestion · {today} ===')

    xls = fetch_workbook(args.url)
    log.info(f'  fetched {len(xls)/1024/1024:.1f} MB')
    rows = parse_workbook(xls)
    log.info(f'  parsed {len(rows)} rows')

    inserted = skipped = 0
    for rec in rows:
        row = build_row(rec, today)
        if not row:
            skipped += 1
            continue
        if args.dry_run:
            log.info(f'    {row["project_name"]} [{row["asset_class"]}] · {row["capacity_mw"]} MW')
            continue
        if upsert_project(client, row):
            inserted += 1
        else:
            skipped += 1

    if not args.dry_run:
        client.table('ingestion_runs').insert({
            'pipeline':           'terna_italy_repowering',
            'status':             'success',
            'started_at':         f'{today}T00:00:00Z',
            'finished_at':        f'{today}T00:00:00Z',
            'records_written':    inserted,
            'source_attribution': f'Terna Anagrafica Impianti ({args.url})',
            'notes':              f'Terna Italy ingestion · {inserted} upserts · {skipped} skipped.',
        }).execute()

    log.info(f'=== complete: {inserted} upserted · {skipped} skipped ===')


if __name__ == '__main__':
    main()
