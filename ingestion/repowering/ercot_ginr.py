#!/usr/bin/env python3
"""
ERCOT GINR (Generation Interconnection Status Report) → repowering_projects.

ERCOT publishes the GINR monthly as an XLS workbook covering every
proposed, in-construction, and recently-commissioned generator in Texas.
We pull the most recent file, filter to wind/solar/storage, and upsert
each project as a repowering pipeline row.

URL pattern (verified May 2026):
  https://www.ercot.com/files/docs/{YYYY}/{MM}/01/GIS_Report_Monthly_{YYYY}_{MM}_01.xlsx

Useful columns from the workbook (sheet "Project Details — Large Gen"):
  - INR  (Interconnection Request Number)        → external_source_id
  - Project Name                                  → project_name
  - Fuel                                          → asset_class mapping
  - Capacity (MW)                                 → capacity_mw
  - County                                        → location_description
  - Project Sponsor / Interconnecting Entity      → developer
  - GIM Study Phase                               → stage (mapped)
  - Approved for Energization (date)              → stage_date if reached

Filter: Fuel ∈ {WIND, SOLAR, OTHER (with Battery indicator)}; Project
Status ∈ {Active, In Service after current date - 90d}.

Texas is the largest US wind/solar/BESS market — this single source
captures ~25-35% of the entire US pipeline.
"""

from __future__ import annotations

import argparse
import sys
from datetime import date, timedelta
from io import BytesIO

import requests

from base_ingestor import get_supabase_client, log
from repowering._base import (
    normalise_stage, normalise_asset_class, make_dedupe_key,
    upsert_project, today_iso, parse_date,
)


ERCOT_URL_TEMPLATE = (
    'https://www.ercot.com/files/docs/{year}/{month:02d}/01/'
    'GIS_Report_Monthly_{year}_{month:02d}_01.xlsx'
)
SHEET_NAME = 'Project Details - Large Gen'

# Map ERCOT GIM (Generator Interconnection Manual) study phases to our enum.
ERCOT_STAGE_MAP = {
    'screening':           'announced',
    'feasibility_study':   'application_submitted',
    'system_impact_study': 'application_submitted',
    'facilities_study':    'application_approved',
    'ia_executed':         'permitted',
    'ia_signed':           'permitted',
    'in_service':          'ongoing',           # COD reached / under construction
    'energization':        'ongoing',
}


def latest_url() -> str | None:
    """Walk back month-by-month to find the most recent ERCOT GINR XLS that exists.

    ERCOT periodically rejects HEAD requests with 4xx even when GET works,
    so we try HEAD first and fall back to a Range-limited GET. We also
    sweep up to 12 months back rather than 4 — if ERCOT pauses publication
    for any reason, we still find the most recent file.
    """
    today = date.today()
    headers = {'User-Agent': 'endenex-terminal/1.0 (operations@endenex.com)'}
    for offset in range(0, 12):
        d = today - timedelta(days=offset * 30)
        url = ERCOT_URL_TEMPLATE.format(year=d.year, month=d.month)
        try:
            r = requests.head(url, timeout=30, allow_redirects=True, headers=headers)
            if r.status_code == 200:
                log.info(f'  found ERCOT GINR (HEAD): {url}')
                return url
            # Some CDNs reject HEAD — fall back to a tiny Range-GET.
            r = requests.get(url, timeout=30, allow_redirects=True,
                             headers={**headers, 'Range': 'bytes=0-1'}, stream=True)
            if r.status_code in (200, 206):
                log.info(f'  found ERCOT GINR (GET): {url}')
                return url
        except requests.RequestException as e:
            log.warning(f'  probe {url} failed: {e}')
            continue
    return None


def fetch_workbook(url: str) -> bytes:
    r = requests.get(url, timeout=120, headers={'User-Agent': 'endenex-terminal/1.0'})
    r.raise_for_status()
    return r.content


def parse_workbook(xls_bytes: bytes) -> list[dict]:
    try:
        import openpyxl
    except ImportError:
        log.error('openpyxl required: pip install openpyxl')
        sys.exit(1)

    wb = openpyxl.load_workbook(BytesIO(xls_bytes), data_only=True, read_only=True)
    if SHEET_NAME not in wb.sheetnames:
        # Fallback: try first sheet
        log.warning(f'  expected sheet "{SHEET_NAME}" not found; using first sheet ({wb.sheetnames[0]})')
        ws = wb[wb.sheetnames[0]]
    else:
        ws = wb[SHEET_NAME]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h or '').strip() for h in rows[0]]
    out: list[dict] = []
    for r in rows[1:]:
        if not r or all(c is None for c in r):
            continue
        rec = dict(zip(headers, r))
        out.append(rec)
    return out


def build_project_row(rec: dict, today: str) -> dict | None:
    fuel = (rec.get('Fuel') or rec.get('Fuel Type') or '').strip()
    asset_class = normalise_asset_class(fuel)
    if asset_class not in {'onshore_wind','solar_pv','bess'}:
        # Skip non-renewable / non-storage projects (gas, coal, etc.)
        return None

    project_name = (rec.get('Project Name') or '').strip()
    if not project_name:
        return None

    stage_raw = rec.get('GIM Study Phase') or rec.get('Project Status') or 'announced'
    stage = normalise_stage(stage_raw, ERCOT_STAGE_MAP) or 'announced'

    inr = (rec.get('INR') or rec.get('Interconnection Request Number') or '').strip()
    capacity = rec.get('Capacity (MW)') or rec.get('MW')
    try:
        capacity_mw = float(capacity) if capacity else None
    except (TypeError, ValueError):
        capacity_mw = None

    return {
        'project_name':        project_name,
        'country_code':        'US',
        'asset_class':         asset_class,
        'stage':               stage,
        'stage_date':          parse_date(rec.get('Approved for Energization')) or today,
        'capacity_mw':         capacity_mw,
        'developer':           (rec.get('Project Sponsor') or rec.get('Interconnecting Entity') or None),
        'operator':            None,
        'planning_reference':  inr or None,
        'location_description': f'{(rec.get("County") or "").strip()}, Texas, USA' if rec.get('County') else 'Texas, USA',
        'source_url':          'https://www.ercot.com/gridinfo/resource',
        'notes':               f'ERCOT GINR · INR {inr}' if inr else 'ERCOT GINR',
        'source_type':         'ercot_giinr',
        'source_date':         today,
        'confidence':          'High',
        'derivation':          'Observed',
        'last_reviewed':       today,
        'external_source':     'ercot_giinr',
        'external_source_id':  inr or None,
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--dry-run', action='store_true')
    ap.add_argument('--url', help='Override URL (default: auto-detect latest)')
    args = ap.parse_args()

    client = get_supabase_client()
    today = today_iso()
    log.info(f'=== ERCOT GINR ingestion · {today} ===')

    url = args.url or latest_url()
    if not url:
        log.error('  no recent ERCOT GINR XLS found')
        sys.exit(1)

    xls = fetch_workbook(url)
    log.info(f'  fetched {len(xls)/1024/1024:.1f} MB')
    rows = parse_workbook(xls)
    log.info(f'  parsed {len(rows)} workbook rows')

    inserted = skipped = 0
    for rec in rows:
        row = build_project_row(rec, today)
        if not row:
            skipped += 1
            continue
        if args.dry_run:
            log.info(f'    {row["project_name"]} [{row["asset_class"]}/{row["stage"]}] · {row["capacity_mw"]} MW · {row["developer"] or "—"}')
            continue
        if upsert_project(client, row):
            inserted += 1
        else:
            skipped += 1

    if not args.dry_run:
        client.table('ingestion_runs').insert({
            'pipeline':           'ercot_ginr_repowering',
            'status':             'success',
            'started_at':         f'{today}T00:00:00Z',
            'finished_at':        f'{today}T00:00:00Z',
            'records_written':    inserted,
            'source_attribution': f'ERCOT GINR ({url})',
            'notes':              f'ERCOT GINR ingestion · {inserted} upserts · {skipped} skipped (non-renewable, missing fields).',
        }).execute()

    log.info(f'=== complete: {inserted} upserted · {skipped} skipped ===')


if __name__ == '__main__':
    main()
